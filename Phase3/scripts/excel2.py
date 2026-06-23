# import json
# from pathlib import Path
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.utils import get_column_letter

# def create_quote_workbook(input_json_path: Path, output_dir: Path):
#     """
#     Creates a new Excel workbook from scratch based on the provided JSON data.
#     """
#     # Load data
#     with open(input_json_path, 'r', encoding='utf-8') as f:
#         data = json.load(f)

#     # Create workbook
#     wb = Workbook()

#     # === Sheet 1: Quote Worksheet ===
#     ws_quote = wb.active
#     ws_quote.title = "Quote Worksheet"

#     # Define cell styles for better formatting
#     header_font = Font(bold=True)
#     header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
#     center_align = Alignment(horizontal="center")

#     # Configuration Section (Top Left)
#     ws_quote['A1'] = "Form Updated"
#     ws_quote['A2'] = "Tax Rate"
#     ws_quote['A3'] = "Markup - Labor"
#     ws_quote['A4'] = "Markup - Other"
#     ws_quote['A5'] = "Lead Brick"
#     ws_quote['A6'] = "Lead Sheet"
#     ws_quote['A7'] = "Steel"
#     ws_quote['A8'] = "Poly 5%"
#     ws_quote['A9'] = "Poly HD"

#     # Set values
#     ws_quote['B1'] = "1/21/25"  # You can make this dynamic if needed
#     ws_quote['B2'] = data["configuration"]["tax_rate"]
#     ws_quote['B3'] = data["configuration"]["markup_labor"]
#     ws_quote['B4'] = data["configuration"]["markup_other"]
#     ws_quote['B5'] = data["configuration"]["lead_brick_price"]
#     ws_quote['B6'] = data["configuration"]["lead_sheet_price"]
#     ws_quote['B7'] = data["configuration"]["steel_price"]
#     ws_quote['B8'] = data["configuration"]["poly_5_price"]
#     ws_quote['B9'] = data["configuration"]["poly_hd_price"]

#     # Format configuration cells
#     for row in range(1, 10):
#         ws_quote.cell(row=row, column=1).font = header_font
#         ws_quote.cell(row=row, column=1).fill = header_fill
#         ws_quote.cell(row=row, column=2).number_format = "0.00%"

#     # Summary Section (Top Right)
#     summary_row = 1
#     summary_headers = ["Cost Total", "Markup %", "Markup", "Burden", "% with Burden", "Tax Amount", "Sale Total"]
#     summary_values = [
#         data["summary"]["cost_total"],
#         data["summary"]["markup_pct"],
#         data["summary"]["markup"],
#         data["summary"]["burden"],
#         data["summary"]["pct_with_burden"],
#         data["summary"]["tax_amount"],
#         data["summary"]["sale_total"]
#     ]

#     for col_idx, (header, value) in enumerate(zip(summary_headers, summary_values), start=11):
#         ws_quote.cell(row=summary_row, column=col_idx).value = header
#         ws_quote.cell(row=summary_row, column=col_idx).font = header_font
#         ws_quote.cell(row=summary_row, column=col_idx).fill = header_fill
#         ws_quote.cell(row=summary_row + 1, column=col_idx).value = value
#         ws_quote.cell(row=summary_row + 1, column=col_idx).number_format = "$#,##0.00"

#     # Item Headers (Row 15)
#     item_headers = [
#         "Qty.", "Unit", "Item", "Description", "Width (inches)", "Length (inches)",
#         "Thickness (inches)", "Weight Per Unit", "Weight Total", "Cost Per Unit",
#         "Cost Total", "Markup %", "Markup", "Burden", "Taxable 1 = Y 0 = N",
#         "Tax Amount", "Total", "Type"
#     ]

#     for col_idx, header in enumerate(item_headers, start=1):
#         ws_quote.cell(row=15, column=col_idx).value = header
#         ws_quote.cell(row=15, column=col_idx).font = header_font
#         ws_quote.cell(row=15, column=col_idx).fill = header_fill
#         ws_quote.cell(row=15, column=col_idx).alignment = center_align

#     # Add Items
#     current_row = 16
#     for item in data["quote_items"]:
#         if item.get("sub_items"):
#             # Write main item row (optional)
#             ws_quote.cell(row=current_row, column=1).value = item["qty"]
#             ws_quote.cell(row=current_row, column=2).value = item["unit"]
#             ws_quote.cell(row=current_row, column=3).value = item["item_name"]
#             ws_quote.cell(row=current_row, column=4).value = item["description"]
#             ws_quote.cell(row=current_row, column=18).value = item["type"]
#             current_row += 1

#             # Write sub-item headers
#             for col_idx, header in enumerate(item_headers[1:], start=2):  # Start from column B
#                 ws_quote.cell(row=current_row, column=col_idx).value = header
#                 ws_quote.cell(row=current_row, column=col_idx).font = Font(italic=True)
#             current_row += 1

#             # Write each sub-item
#             for sub in item["sub_items"]:
#                 ws_quote.cell(row=current_row, column=1).value = sub["qty"]
#                 ws_quote.cell(row=current_row, column=2).value = sub["unit"]
#                 ws_quote.cell(row=current_row, column=3).value = sub["item_name"]
#                 ws_quote.cell(row=current_row, column=4).value = sub["description"]
#                 ws_quote.cell(row=current_row, column=5).value = sub["width"]
#                 ws_quote.cell(row=current_row, column=6).value = sub["length"]
#                 ws_quote.cell(row=current_row, column=7).value = sub["thickness"]
#                 ws_quote.cell(row=current_row, column=8).value = sub["weight_per_unit"]
#                 ws_quote.cell(row=current_row, column=9).value = sub["weight_total"]
#                 ws_quote.cell(row=current_row, column=10).value = sub["cost_per_unit"]
#                 ws_quote.cell(row=current_row, column=11).value = sub["cost_total"]
#                 ws_quote.cell(row=current_row, column=12).value = sub["markup_pct"]
#                 ws_quote.cell(row=current_row, column=13).value = sub["markup"]
#                 ws_quote.cell(row=current_row, column=14).value = sub["burden"]
#                 ws_quote.cell(row=current_row, column=15).value = sub["taxable"]
#                 ws_quote.cell(row=current_row, column=16).value = sub["tax_amount"]
#                 ws_quote.cell(row=current_row, column=17).value = sub["total"]
#                 ws_quote.cell(row=current_row, column=18).value = sub["type"]
#                 current_row += 1

#             # Add blank row after assembly
#             current_row += 1
#         else:
#             # Simple item
#             ws_quote.cell(row=current_row, column=1).value = item["qty"]
#             ws_quote.cell(row=current_row, column=2).value = item["unit"]
#             ws_quote.cell(row=current_row, column=3).value = item["item_name"]
#             ws_quote.cell(row=current_row, column=4).value = item["description"]
#             ws_quote.cell(row=current_row, column=5).value = item["width"]
#             ws_quote.cell(row=current_row, column=6).value = item["length"]
#             ws_quote.cell(row=current_row, column=7).value = item["thickness"]
#             ws_quote.cell(row=current_row, column=8).value = item["weight_per_unit"]
#             ws_quote.cell(row=current_row, column=9).value = item["weight_total"]
#             ws_quote.cell(row=current_row, column=10).value = item["cost_per_unit"]
#             ws_quote.cell(row=current_row, column=11).value = item["cost_total"]
#             ws_quote.cell(row=current_row, column=12).value = item["markup_pct"]
#             ws_quote.cell(row=current_row, column=13).value = item["markup"]
#             ws_quote.cell(row=current_row, column=14).value = item["burden"]
#             ws_quote.cell(row=current_row, column=15).value = item["taxable"]
#             ws_quote.cell(row=current_row, column=16).value = item["tax_amount"]
#             ws_quote.cell(row=current_row, column=17).value = item["total"]
#             ws_quote.cell(row=current_row, column=18).value = item["type"]
#             current_row += 1

#     # Freeze top 15 rows for easy scrolling
#     ws_quote.freeze_panes = 'A16'

#     # Auto-fit columns
#     for col in ws_quote.columns:
#         max_length = 0
#         column = col[0].column_letter
#         for cell in col:
#             try:
#                 if len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
#         adjusted_width = (max_length + 2)
#         ws_quote.column_dimensions[column].width = adjusted_width

#     # === Sheet 2: Items ===
#     ws_items = wb.create_sheet("Items")
#     item_headers = ["Item", "Description", "Cost", "Burden", "U/M", "Width", "Length", "Thickness", "Weight Per Unit", "Type"]
#     for col_idx, header in enumerate(item_headers, start=1):
#         ws_items.cell(row=1, column=col_idx).value = header
#         ws_items.cell(row=1, column=col_idx).font = header_font
#         ws_items.cell(row=1, column=col_idx).fill = header_fill
#         ws_items.cell(row=1, column=col_idx).alignment = center_align

#     for row_idx, item in enumerate(data["items"], start=2):
#         ws_items.cell(row=row_idx, column=1).value = item["item"]
#         ws_items.cell(row=row_idx, column=2).value = item["description"]
#         ws_items.cell(row=row_idx, column=3).value = item["cost"]
#         ws_items.cell(row=row_idx, column=4).value = item["burden"]
#         ws_items.cell(row=row_idx, column=5).value = item["uom"]
#         ws_items.cell(row=row_idx, column=6).value = item["width"]
#         ws_items.cell(row=row_idx, column=7).value = item["length"]
#         ws_items.cell(row=row_idx, column=8).value = item["thickness"]
#         ws_items.cell(row=row_idx, column=9).value = item["weight_per_unit"]
#         ws_items.cell(row=row_idx, column=10).value = item["type"]

#     # Auto-fit columns for Items sheet
#     for col in ws_items.columns:
#         max_length = 0
#         column = col[0].column_letter
#         for cell in col:
#             try:
#                 if len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
#         adjusted_width = (max_length + 2)
#         ws_items.column_dimensions[column].width = adjusted_width

#     # === Sheet 3: GroupItems ===
#     ws_group = wb.create_sheet("GroupItems")
#     ws_group.append(["Group Item"])
#     ws_group.cell(row=1, column=1).font = header_font
#     ws_group.cell(row=1, column=1).fill = header_fill
#     for name in data["group_items"]:
#         ws_group.append([name])

#     # === Sheet 4: ItemTemplates ===
#     ws_templates = wb.create_sheet("ItemTemplates")
#     ws_templates.append(["Template Name", "Type", "Component Items"])
#     ws_templates.cell(row=1, column=1).font = header_font
#     ws_templates.cell(row=1, column=1).fill = header_fill
#     ws_templates.cell(row=1, column=2).font = header_font
#     ws_templates.cell(row=1, column=2).fill = header_fill
#     ws_templates.cell(row=1, column=3).font = header_font
#     ws_templates.cell(row=1, column=3).fill = header_fill

#     for tmpl in data["item_templates"]:
#         comp_str = "; ".join([c["item"] for c in tmpl["components"]])
#         ws_templates.append([tmpl["template_name"], tmpl["type"], comp_str])

#     # === Sheet 5: LeadTimes ===
#     ws_lead = wb.create_sheet("LeadTimes")
#     ws_lead.append(["Task", "Hours", "Personnel"])
#     ws_lead.cell(row=1, column=1).font = header_font
#     ws_lead.cell(row=1, column=1).fill = header_fill
#     ws_lead.cell(row=1, column=2).font = header_font
#     ws_lead.cell(row=1, column=2).fill = header_fill
#     ws_lead.cell(row=1, column=3).font = header_font
#     ws_lead.cell(row=1, column=3).fill = header_fill

#     for lt in data["lead_times"]:
#         ws_lead.append([lt["task"], lt["hours"], lt["personnel"]])

#     # Save the workbook
#     timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
#     output_file = output_dir / f"Generated_Quote_{timestamp}.xlsx"
#     wb.save(output_file)
#     print(f"✅ New workbook created: {output_file}")
#     print(f"📊 Total items added: {len(data['quote_items'])}")

# # === Main Execution ===
# if __name__ == "__main__":
#     from datetime import datetime

#     SCRIPT_DIR = Path(__file__).parent
#     ROOT_DIR = SCRIPT_DIR.parent
#     INPUT_DIR = ROOT_DIR / "input"
#     OUTPUT_DIR = ROOT_DIR / "output"

#     # Ensure output directory exists
#     OUTPUT_DIR.mkdir(exist_ok=True)

#     # Use the first JSON file found in input directory
#     input_files = list(INPUT_DIR.glob("*.json"))
#     if not input_files:
#         print(f"❌ No JSON files found in {INPUT_DIR}")
#         exit(1)

#     input_json_path = input_files[0]
#     print(f"📄 Using input file: {input_json_path}")

#     create_quote_workbook(input_json_path, OUTPUT_DIR)


import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import sys

def create_quote_workbook(input_json_path: Path, output_dir: Path):
    """
    Creates a new Excel workbook from scratch based on the provided JSON data.
    """
    # Load data
    with open(input_json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Create workbook
    wb = Workbook()

    # === Sheet 1: Quote Worksheet ===
    ws_quote = wb.active
    ws_quote.title = "Quote Worksheet"

    # Define cell styles for better formatting
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    center_align = Alignment(horizontal="center")

    # Configuration Section (Top Left)
    ws_quote['A1'] = "Form Updated"
    ws_quote['A2'] = "Tax Rate"
    ws_quote['A3'] = "Markup - Labor"
    ws_quote['A4'] = "Markup - Other"
    ws_quote['A5'] = "Lead Brick"
    ws_quote['A6'] = "Lead Sheet"
    ws_quote['A7'] = "Steel"
    ws_quote['A8'] = "Poly 5%"
    ws_quote['A9'] = "Poly HD"

    # Set values
    config = data.get("configuration", {})
    ws_quote['B1'] = "1/21/25"  # You can make this dynamic if needed
    ws_quote['B2'] = config.get("tax_rate", 0.0825)
    ws_quote['B3'] = config.get("markup_labor", 0.0)
    ws_quote['B4'] = config.get("markup_other", 0.25)
    ws_quote['B5'] = config.get("lead_brick_price", 1.25)
    ws_quote['B6'] = config.get("lead_sheet_price", 1.75)
    ws_quote['B7'] = config.get("steel_price", 1.40)
    ws_quote['B8'] = config.get("poly_5_price", 850.00)
    ws_quote['B9'] = config.get("poly_hd_price", 350.00)

    # Format configuration cells
    for row in range(1, 10):
        ws_quote.cell(row=row, column=1).font = header_font
        ws_quote.cell(row=row, column=1).fill = header_fill
        ws_quote.cell(row=row, column=2).number_format = "0.00%"

    # Summary Section (Top Right)
    # summary_row = 1
    # summary_headers = ["Cost Total", "Markup %", "Markup", "Burden", "% with Burden", "Tax Amount", "Sale Total"]
    # summary_values = [
    #     data["summary"].get("cost_total", 0),
    #     data["summary"].get("markup_pct", 0),
    #     data["summary"].get("markup", 0),
    #     data["summary"].get("burden", 0),
    #     data["summary"].get("pct_with_burden", 0),
    #     data["summary"].get("tax_amount", 0),
    #     data["summary"].get("sale_total", 0)
    # ]

    # for col_idx, (header, value) in enumerate(summary_headers, start=11):
    #     ws_quote.cell(row=summary_row, column=col_idx).value = header
    #     ws_quote.cell(row=summary_row, column=col_idx).font = header_font
    #     ws_quote.cell(row=summary_row, column=col_idx).fill = header_fill
    #     ws_quote.cell(row=summary_row + 1, column=col_idx).value = value
    #     ws_quote.cell(row=summary_row + 1, column=col_idx).number_format = "$#,##0.00"
    # Summary Section (Top Right)
    summary_row = 1
    summary_headers = ["Cost Total", "Markup %", "Markup", "Burden", "% with Burden", "Tax Amount", "Sale Total"]
    summary_values = [
        data["summary"].get("cost_total", 0),
        data["summary"].get("markup_pct", 0),
        data["summary"].get("markup", 0),
        data["summary"].get("burden", 0),
        data["summary"].get("pct_with_burden", 0),
        data["summary"].get("tax_amount", 0),
        data["summary"].get("sale_total", 0)
    ]

    # Zip headers and values together
    for col_idx, (header, value) in enumerate(zip(summary_headers, summary_values), start=11):
        ws_quote.cell(row=summary_row, column=col_idx).value = header
        ws_quote.cell(row=summary_row, column=col_idx).font = header_font
        ws_quote.cell(row=summary_row, column=col_idx).fill = header_fill
        ws_quote.cell(row=summary_row + 1, column=col_idx).value = value
        ws_quote.cell(row=summary_row + 1, column=col_idx).number_format = "$#,##0.00"

    # Item Headers (Row 15)
    item_headers = [
        "Qty.", "Unit", "Item", "Description", "Width (inches)", "Length (inches)",
        "Thickness (inches)", "Weight Per Unit", "Weight Total", "Cost Per Unit",
        "Cost Total", "Markup %", "Markup", "Burden", "Taxable 1 = Y 0 = N",
        "Tax Amount", "Total", "Type"
    ]

    for col_idx, header in enumerate(item_headers, start=1):
        ws_quote.cell(row=15, column=col_idx).value = header
        ws_quote.cell(row=15, column=col_idx).font = header_font
        ws_quote.cell(row=15, column=col_idx).fill = header_fill
        ws_quote.cell(row=15, column=col_idx).alignment = center_align

    # Add Items
    current_row = 16
    quote_items = data.get("quote_items", [])
    for item in quote_items:
        if item.get("sub_items"):
            # Write main item row (optional)
            ws_quote.cell(row=current_row, column=1).value = item["qty"]
            ws_quote.cell(row=current_row, column=2).value = item["unit"]
            ws_quote.cell(row=current_row, column=3).value = item["item_name"]
            ws_quote.cell(row=current_row, column=4).value = item["description"]
            ws_quote.cell(row=current_row, column=18).value = item["type"]
            current_row += 1

            # Write sub-item headers
            for col_idx, header in enumerate(item_headers[1:], start=2):  # Start from column B
                ws_quote.cell(row=current_row, column=col_idx).value = header
                ws_quote.cell(row=current_row, column=col_idx).font = Font(italic=True)
            current_row += 1

            # Write each sub-item
            for sub in item["sub_items"]:
                ws_quote.cell(row=current_row, column=1).value = sub["qty"]
                ws_quote.cell(row=current_row, column=2).value = sub["unit"]
                ws_quote.cell(row=current_row, column=3).value = sub["item_name"]
                ws_quote.cell(row=current_row, column=4).value = sub["description"]
                ws_quote.cell(row=current_row, column=5).value = sub["width"]
                ws_quote.cell(row=current_row, column=6).value = sub["length"]
                ws_quote.cell(row=current_row, column=7).value = sub["thickness"]
                ws_quote.cell(row=current_row, column=8).value = sub["weight_per_unit"]
                ws_quote.cell(row=current_row, column=9).value = sub["weight_total"]
                ws_quote.cell(row=current_row, column=10).value = sub["cost_per_unit"]
                ws_quote.cell(row=current_row, column=11).value = sub["cost_total"]
                ws_quote.cell(row=current_row, column=12).value = sub["markup_pct"]
                ws_quote.cell(row=current_row, column=13).value = sub["markup"]
                ws_quote.cell(row=current_row, column=14).value = sub["burden"]
                ws_quote.cell(row=current_row, column=15).value = sub["taxable"]
                ws_quote.cell(row=current_row, column=16).value = sub["tax_amount"]
                ws_quote.cell(row=current_row, column=17).value = sub["total"]
                ws_quote.cell(row=current_row, column=18).value = sub["type"]
                current_row += 1

            # Add blank row after assembly
            current_row += 1
        else:
            # Simple item
            ws_quote.cell(row=current_row, column=1).value = item["qty"]
            ws_quote.cell(row=current_row, column=2).value = item["unit"]
            ws_quote.cell(row=current_row, column=3).value = item["item_name"]
            ws_quote.cell(row=current_row, column=4).value = item["description"]
            ws_quote.cell(row=current_row, column=5).value = item["width"]
            ws_quote.cell(row=current_row, column=6).value = item["length"]
            ws_quote.cell(row=current_row, column=7).value = item["thickness"]
            ws_quote.cell(row=current_row, column=8).value = item["weight_per_unit"]
            ws_quote.cell(row=current_row, column=9).value = item["weight_total"]
            ws_quote.cell(row=current_row, column=10).value = item["cost_per_unit"]
            ws_quote.cell(row=current_row, column=11).value = item["cost_total"]
            ws_quote.cell(row=current_row, column=12).value = item["markup_pct"]
            ws_quote.cell(row=current_row, column=13).value = item["markup"]
            ws_quote.cell(row=current_row, column=14).value = item["burden"]
            ws_quote.cell(row=current_row, column=15).value = item["taxable"]
            ws_quote.cell(row=current_row, column=16).value = item["tax_amount"]
            ws_quote.cell(row=current_row, column=17).value = item["total"]
            ws_quote.cell(row=current_row, column=18).value = item["type"]
            current_row += 1

    # Freeze top 15 rows for easy scrolling
    ws_quote.freeze_panes = 'A16'

    # Auto-fit columns
    for col in ws_quote.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_quote.column_dimensions[column].width = adjusted_width

    # === Sheet 2: Items ===
    ws_items = wb.create_sheet("Items")
    item_headers = ["Item", "Description", "Cost", "Burden", "U/M", "Width", "Length", "Thickness", "Weight Per Unit", "Type"]
    for col_idx, header in enumerate(item_headers, start=1):
        ws_items.cell(row=1, column=col_idx).value = header
        ws_items.cell(row=1, column=col_idx).font = header_font
        ws_items.cell(row=1, column=col_idx).fill = header_fill
        ws_items.cell(row=1, column=col_idx).alignment = center_align

    items = data.get("items", [])
    for row_idx, item in enumerate(items, start=2):
        ws_items.cell(row=row_idx, column=1).value = item["item"]
        ws_items.cell(row=row_idx, column=2).value = item["description"]
        ws_items.cell(row=row_idx, column=3).value = item.get("cost", 0)
        ws_items.cell(row=row_idx, column=4).value = item.get("burden", 0)
        ws_items.cell(row=row_idx, column=5).value = item.get("uom", "")
        ws_items.cell(row=row_idx, column=6).value = item.get("width", 0)
        ws_items.cell(row=row_idx, column=7).value = item.get("length", 0)
        ws_items.cell(row=row_idx, column=8).value = item.get("thickness", 0)
        ws_items.cell(row=row_idx, column=9).value = item.get("weight_per_unit", 0)
        ws_items.cell(row=row_idx, column=10).value = item.get("type", "")

    # Auto-fit columns for Items sheet
    for col in ws_items.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_items.column_dimensions[column].width = adjusted_width

    # === Sheet 3: GroupItems ===
    ws_group = wb.create_sheet("GroupItems")
    ws_group.append(["Group Item"])
    ws_group.cell(row=1, column=1).font = header_font
    ws_group.cell(row=1, column=1).fill = header_fill
    group_items = data.get("group_items", [])
    for name in group_items:
        ws_group.append([name])

    # === Sheet 4: ItemTemplates ===
    ws_templates = wb.create_sheet("ItemTemplates")
    ws_templates.append(["Template Name", "Type", "Component Items"])
    ws_templates.cell(row=1, column=1).font = header_font
    ws_templates.cell(row=1, column=1).fill = header_fill
    ws_templates.cell(row=1, column=2).font = header_font
    ws_templates.cell(row=1, column=2).fill = header_fill
    ws_templates.cell(row=1, column=3).font = header_font
    ws_templates.cell(row=1, column=3).fill = header_fill

    item_templates = data.get("item_templates", [])
    for tmpl in item_templates:
        comp_str = "; ".join([c["item"] for c in tmpl.get("components", [])])
        ws_templates.append([tmpl["template_name"], tmpl["type"], comp_str])

    # === Sheet 5: LeadTimes ===
    ws_lead = wb.create_sheet("LeadTimes")
    ws_lead.append(["Task", "Hours", "Personnel"])
    ws_lead.cell(row=1, column=1).font = header_font
    ws_lead.cell(row=1, column=1).fill = header_fill
    ws_lead.cell(row=1, column=2).font = header_font
    ws_lead.cell(row=1, column=2).fill = header_fill
    ws_lead.cell(row=1, column=3).font = header_font
    ws_lead.cell(row=1, column=3).fill = header_fill

    lead_times = data.get("lead_times", [])
    for lt in lead_times:
        ws_lead.append([lt["task"], lt["hours"], lt["personnel"]])

    # Save the workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = output_dir / f"Generated_Quote_{timestamp}.xlsx"
    wb.save(output_file)
    print(f"✅ New workbook created: {output_file}")
    print(f"📊 Total items added: {len(quote_items)}")

# === Main Execution ===
if __name__ == "__main__":
    from datetime import datetime

    SCRIPT_DIR = Path(__file__).parent
    ROOT_DIR = SCRIPT_DIR.parent
    INPUT_DIR = ROOT_DIR / "input"
    OUTPUT_DIR = ROOT_DIR / "output"

    # Ensure output directory exists
    OUTPUT_DIR.mkdir(exist_ok=True)

    # --- Get input file name ---
    # Option 1: Use command-line argument if provided
    if len(sys.argv) > 1:
        input_filename = sys.argv[1]
        input_json_path = INPUT_DIR / input_filename
    else:
        # Option 2: Use a default filename (e.g., "bid_data.json")
        input_filename = "quote_data2.json"
        input_json_path = INPUT_DIR / input_filename

    # Check if file exists
    if not input_json_path.exists():
        print(f"❌ Input file not found: {input_json_path}")
        exit(1)

    print(f"📄 Using input file: {input_json_path}")

    create_quote_workbook(input_json_path, OUTPUT_DIR)