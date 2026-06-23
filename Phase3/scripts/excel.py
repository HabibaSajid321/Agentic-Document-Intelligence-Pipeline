# # scripts/fill_quote_worksheet.py
# import openpyxl
# from pathlib import Path
# import json
# from datetime import datetime

# def fill_quote_worksheet(data):
#     """
#     Fill the Quote Worksheet Excel file with demo data.
#     """
#     SCRIPT_DIR = Path(__file__).parent
#     ROOT_DIR = SCRIPT_DIR.parent
#     TEMPLATE_DIR = ROOT_DIR / "templates"
#     OUTPUT_DIR = ROOT_DIR / "output"
    
#     TEMPLATE_FILE = TEMPLATE_DIR / "worksheet.xlsm"
#     OUTPUT_FILE = OUTPUT_DIR / f"Filled_Quote.xlsx"
    
#     OUTPUT_DIR.mkdir(exist_ok=True)
    
#     if not TEMPLATE_FILE.exists():
#         print(f"❌ Template file not found: {TEMPLATE_FILE}")
#         return
    
#     # Load workbook (keep_vba=True to preserve macros, but save as xlsx)
#     wb = openpyxl.load_workbook(TEMPLATE_FILE, keep_vba=False)
#     ws = wb.active  # Assuming first sheet is the quote worksheet
    
#     print("✅ Loaded Excel template")
#     print(f"📄 Active Sheet: {ws.title}")
    
#     # Update configuration section
#     config = data.get("configuration", {})
#     ws['D1'] = config.get("tax_rate", 0.00)  # Tax Rate
#     ws['D2'] = config.get("markup_labor", 0.00)  # Markup - Labor
#     ws['D3'] = config.get("markup_other", 0.25)  # Markup - Other
#     ws['D4'] = config.get("lead_brick_price", 1.25)  # Lead Brick
#     ws['D5'] = config.get("lead_sheet_price", 1.75)  # Lead Sheet
#     ws['D6'] = config.get("steel_price", 1.40)  # Steel
#     ws['D7'] = config.get("poly_5_price", 850.00)  # Poly 5%
#     ws['D8'] = config.get("poly_hd_price", 350.00)  # Poly HD
    
#     print("\n✓ Configuration updated")
    
#     # Get quote items
#     quote_items = data.get("quote_items", [])
    
#     # Starting row for items (after headers)
#     current_row = 20  # Starting after the first example item
    
#     for idx, item in enumerate(quote_items, 1):
#         print(f"\n📦 Adding Item {idx}: {item.get('item_name', 'Unnamed')}")
        
#         # Main item header row
#         ws[f'A{current_row}'] = item.get('qty', 1)
#         ws[f'B{current_row}'] = item.get('unit', 'ea')
#         ws[f'C{current_row}'] = item.get('item_name', '')
#         ws[f'D{current_row}'] = item.get('description', '')
#         ws[f'E{current_row}'] = item.get('width', 0.0)
#         ws[f'F{current_row}'] = item.get('length', 0.0)
#         ws[f'G{current_row}'] = item.get('thickness', 0.0)
#         ws[f'H{current_row}'] = item.get('weight_per_unit', 0.0)
#         ws[f'I{current_row}'] = item.get('weight_total', 0.0)
#         ws[f'J{current_row}'] = item.get('cost_per_unit', 0.0)
#         ws[f'K{current_row}'] = item.get('cost_total', 0.0)
#         ws[f'L{current_row}'] = item.get('markup_pct', 0.0)
#         ws[f'M{current_row}'] = item.get('markup', 0.0)
#         ws[f'N{current_row}'] = item.get('burden', 0.0)
#         ws[f'O{current_row}'] = item.get('taxable', 0)
#         ws[f'P{current_row}'] = item.get('tax_amount', 0.0)
#         ws[f'Q{current_row}'] = item.get('total', 0.0)
#         ws[f'R{current_row}'] = item.get('type', 'Material')
        
#         current_row += 1
        
#         # Add sub-items (labor, materials, etc.)
#         sub_items = item.get('sub_items', [])
#         for sub in sub_items:
#             ws[f'A{current_row}'] = sub.get('qty', 0)
#             ws[f'B{current_row}'] = sub.get('unit', 'ea')
#             ws[f'C{current_row}'] = sub.get('item_name', '')
#             ws[f'D{current_row}'] = sub.get('description', '')
#             ws[f'E{current_row}'] = sub.get('width', 0.0)
#             ws[f'F{current_row}'] = sub.get('length', 0.0)
#             ws[f'G{current_row}'] = sub.get('thickness', 0.0)
#             ws[f'H{current_row}'] = ''
#             ws[f'I{current_row}'] = sub.get('weight_total', 0.0)
#             ws[f'J{current_row}'] = sub.get('cost_per_unit', 0.0)
#             ws[f'K{current_row}'] = sub.get('cost_total', 0.0)
#             ws[f'L{current_row}'] = sub.get('markup_pct', 0.0)
#             ws[f'M{current_row}'] = sub.get('markup', 0.0)
#             ws[f'N{current_row}'] = sub.get('burden', 0.0)
#             ws[f'O{current_row}'] = sub.get('taxable', 0)
#             ws[f'P{current_row}'] = sub.get('tax_amount', 0.0)
#             ws[f'Q{current_row}'] = sub.get('total', 0.0)
#             ws[f'R{current_row}'] = sub.get('type', 'Material')
            
#             print(f"  └─ {sub.get('item_name', 'Unnamed sub-item')}: ${sub.get('total', 0.0):.2f}")
#             current_row += 1
        
#         # Add blank row between items
#         current_row += 1
    
#     # Update summary cells if provided
#     summary = data.get("summary", {})
#     if summary:
#         # These cell references may need adjustment based on actual template
#         ws['K2'] = summary.get('cost_total', 0.0)
#         ws['L2'] = summary.get('markup_pct', 0.27)
#         ws['M2'] = summary.get('markup', 0.0)
#         ws['N2'] = summary.get('burden', 0.0)
#         ws['O2'] = summary.get('pct_with_burden', 0.36)
#         ws['P2'] = summary.get('tax_amount', 0.0)
#         ws['Q2'] = summary.get('sale_total', 0.0)
    
#     # Save the file
#     wb.save(OUTPUT_FILE)
#     print(f"\n✅ Quote worksheet saved to: {OUTPUT_FILE}")
#     print(f"📊 Total items added: {len(quote_items)}")

# def main():
#     # Load demo data
#     SCRIPT_DIR = Path(__file__).parent
#     ROOT_DIR = SCRIPT_DIR.parent
#     INPUT_DIR = ROOT_DIR / "input"
    
#     DEMO_DATA_FILE = INPUT_DIR / "quote_data.json"
    
#     if not DEMO_DATA_FILE.exists():
#         print(f"❌ Demo data file not found: {DEMO_DATA_FILE}")
#         print("Creating sample data file...")
        
#         # Create sample data structure
#         sample_data = {
#             "configuration": {
#                 "tax_rate": 0.0825,
#                 "markup_labor": 0.00,
#                 "markup_other": 0.25,
#                 "lead_brick_price": 1.25,
#                 "lead_sheet_price": 1.75,
#                 "steel_price": 1.40,
#                 "poly_5_price": 850.00,
#                 "poly_hd_price": 350.00
#             },
#             "quote_items": [
#                 {
#                     "qty": 1,
#                     "unit": "ea",
#                     "item_name": "4\" x 12\" x 2\" Lead Brick",
#                     "description": "",
#                     "width": 4.0,
#                     "length": 12.0,
#                     "thickness": 2.0,
#                     "weight_per_unit": 40.0,
#                     "weight_total": 40.0,
#                     "cost_per_unit": 0.0,
#                     "cost_total": 0.0,
#                     "markup_pct": 0.0,
#                     "markup": 0.0,
#                     "burden": 0.0,
#                     "taxable": 0,
#                     "tax_amount": 0.0,
#                     "total": 84.0,
#                     "type": "Assembly",
#                     "sub_items": [
#                         {
#                             "qty": 0.134,
#                             "unit": "mh",
#                             "item_name": "Labor - Laborer - ST",
#                             "description": "Labor Fabricate - Brick",
#                             "width": 0.0,
#                             "length": 0.0,
#                             "thickness": 0.0,
#                             "weight_total": 0.0,
#                             "cost_per_unit": 40.00,
#                             "cost_total": 5.36,
#                             "markup_pct": 0.00,
#                             "markup": 0.00,
#                             "burden": 5.49,
#                             "taxable": 0,
#                             "tax_amount": 0.0,
#                             "total": 11.00,
#                             "type": "Labor"
#                         },
#                         {
#                             "qty": 1,
#                             "unit": "ea",
#                             "item_name": "Gases - Pro-Fuel",
#                             "description": "Pro Fuel",
#                             "width": 0.0,
#                             "length": 0.0,
#                             "thickness": 0.0,
#                             "weight_total": 0.0,
#                             "cost_per_unit": 0.17,
#                             "cost_total": 0.17,
#                             "markup_pct": 0.25,
#                             "markup": 0.04,
#                             "burden": 0.0,
#                             "taxable": 0,
#                             "tax_amount": 0.0,
#                             "total": 1.00,
#                             "type": "Material"
#                         },
#                         {
#                             "qty": 1,
#                             "unit": "ea",
#                             "item_name": "Gases - Natural Gas",
#                             "description": "Natural Gas",
#                             "width": 0.0,
#                             "length": 0.0,
#                             "thickness": 0.0,
#                             "weight_total": 0.0,
#                             "cost_per_unit": 0.23,
#                             "cost_total": 0.23,
#                             "markup_pct": 0.25,
#                             "markup": 0.06,
#                             "burden": 0.0,
#                             "taxable": 0,
#                             "tax_amount": 0.0,
#                             "total": 1.00,
#                             "type": "Material"
#                         },
#                         {
#                             "qty": 1,
#                             "unit": "ea",
#                             "item_name": "Lead - Ingot",
#                             "description": "Lead Ingot for Brick",
#                             "width": 4.0,
#                             "length": 12.0,
#                             "thickness": 2.0,
#                             "weight_total": 40.0,
#                             "cost_per_unit": 56.00,
#                             "cost_total": 56.00,
#                             "markup_pct": 0.25,
#                             "markup": 14.00,
#                             "burden": 0.0,
#                             "taxable": 0,
#                             "tax_amount": 0.0,
#                             "total": 71.00,
#                             "type": "Material"
#                         }
#                     ]
#                 },
#                 {
#                     "qty": 100,
#                     "unit": "sf",
#                     "item_name": "Lead Sheet 1/8\" Thick",
#                     "description": "Wall Shielding",
#                     "width": 48.0,
#                     "length": 96.0,
#                     "thickness": 0.125,
#                     "weight_per_unit": 5.2,
#                     "weight_total": 520.0,
#                     "cost_per_unit": 9.10,
#                     "cost_total": 910.00,
#                     "markup_pct": 0.25,
#                     "markup": 227.50,
#                     "burden": 0.0,
#                     "taxable": 0,
#                     "tax_amount": 0.0,
#                     "total": 1137.50,
#                     "type": "Material",
#                     "sub_items": []
#                 }
#             ],
#             "summary": {
#                 "cost_total": 61.77,
#                 "markup_pct": 0.27,
#                 "markup": 16.74,
#                 "burden": 5.49,
#                 "pct_with_burden": 0.36,
#                 "tax_amount": 0.0,
#                 "sale_total": 84.00
#             }
#         }
        
#         INPUT_DIR.mkdir(exist_ok=True)
#         with open(DEMO_DATA_FILE, 'w', encoding='utf-8') as f:
#             json.dump(sample_data, f, indent=2)
        
#         print(f"✅ Sample data created: {DEMO_DATA_FILE}")
    
#     # Load data
#     with open(DEMO_DATA_FILE, 'r', encoding='utf-8') as f:
#         data = json.load(f)
    
#     print(f"\n{'='*80}")
#     print("FILLING QUOTE WORKSHEET")
#     print('='*80)
    
#     fill_quote_worksheet(data)

# if __name__ == "__main__":
#     main()

                                #  currently working code
# import openpyxl
# from pathlib import Path
# import json
# from datetime import datetime

# def fill_quote_worksheet(data):
#     """
#     Fill the Quote Worksheet Excel file with demo data.
#     """
#     SCRIPT_DIR = Path(__file__).parent
#     ROOT_DIR = SCRIPT_DIR.parent
#     TEMPLATE_DIR = ROOT_DIR / "templates"
#     OUTPUT_DIR = ROOT_DIR / "output"
    
#     TEMPLATE_FILE = TEMPLATE_DIR / "worksheet.xlsm"
#     OUTPUT_FILE = OUTPUT_DIR / f"Filled_Quote.xlsx"
    
#     OUTPUT_DIR.mkdir(exist_ok=True)
    
#     if not TEMPLATE_FILE.exists():
#         print(f"❌ Template file not found: {TEMPLATE_FILE}")
#         return
    
#     # Load workbook
#     wb = openpyxl.load_workbook(TEMPLATE_FILE, keep_vba=False)
    
#     # CRITICAL FIX: Select the correct sheet by name
#     if "Quote Worksheet" in wb.sheetnames:
#         ws = wb["Quote Worksheet"]
#     else:
#         print(f"⚠️  'Quote Worksheet' not found. Available sheets: {wb.sheetnames}")
#         ws = wb.active
    
#     print("✅ Loaded Excel template")
#     print(f"📄 Active Sheet: {ws.title}")
    
#     # Update configuration section (D1-D8)
#     config = data.get("configuration", {})
#     ws['D1'] = config.get("tax_rate", 0.00)  # Tax Rate
#     ws['D2'] = config.get("markup_labor", 0.00)  # Markup - Labor
#     ws['D3'] = config.get("markup_other", 0.25)  # Markup - Other
#     ws['D4'] = config.get("lead_brick_price", 1.25)  # Lead Brick
#     ws['D5'] = config.get("lead_sheet_price", 1.75)  # Lead Sheet
#     ws['D6'] = config.get("steel_price", 1.40)  # Steel
#     ws['D7'] = config.get("poly_5_price", 850.00)  # Poly 5%
#     ws['D8'] = config.get("poly_hd_price", 350.00)  # Poly HD
    
#     print("\n✓ Configuration updated")
    
#     # Get quote items
#     quote_items = data.get("quote_items", [])
    
#     # CRITICAL FIX: Find the first empty row after existing data
#     # Start searching from row 24 (after the example item)
#     current_row = 24
#     while ws[f'A{current_row}'].value is not None or ws[f'C{current_row}'].value is not None:
#         current_row += 1
#         if current_row > 100:  # Safety limit
#             break
    
#     print(f"\n🔍 Starting insertion at row: {current_row}")
    
#     for idx, item in enumerate(quote_items, 1):
#         print(f"\n📦 Adding Item {idx}: {item.get('item_name', 'Unnamed')}")
        
#         # Main item header row
#         ws[f'A{current_row}'] = item.get('qty', 1)
#         ws[f'B{current_row}'] = item.get('unit', 'ea')
#         ws[f'C{current_row}'] = item.get('item_name', '')
#         ws[f'D{current_row}'] = item.get('description', '')
#         ws[f'E{current_row}'] = item.get('width', 0.0)
#         ws[f'F{current_row}'] = item.get('length', 0.0)
#         ws[f'G{current_row}'] = item.get('thickness', 0.0)
#         ws[f'H{current_row}'] = item.get('weight_per_unit', 0.0) if item.get('weight_per_unit') else ''
#         ws[f'I{current_row}'] = item.get('weight_total', 0.0)
#         ws[f'J{current_row}'] = item.get('cost_per_unit', 0.0)
#         ws[f'K{current_row}'] = item.get('cost_total', 0.0)
#         ws[f'L{current_row}'] = item.get('markup_pct', 0.0)
#         ws[f'M{current_row}'] = item.get('markup', 0.0)
#         ws[f'N{current_row}'] = item.get('burden', 0.0)
#         ws[f'O{current_row}'] = item.get('taxable', 0)
#         ws[f'P{current_row}'] = item.get('tax_amount', 0.0)
#         ws[f'Q{current_row}'] = item.get('total', 0.0)
#         ws[f'R{current_row}'] = item.get('type', 'Material')
        
#         current_row += 1
        
#         # Add sub-items (labor, materials, etc.)
#         sub_items = item.get('sub_items', [])
#         if sub_items:
#             # Add header row for sub-items (columns labels)
#             ws[f'A{current_row}'] = 'Qty.'
#             ws[f'B{current_row}'] = 'Unit'
#             ws[f'C{current_row}'] = 'Item'
#             ws[f'D{current_row}'] = 'Description'
#             ws[f'E{current_row}'] = 'Width (inches)'
#             ws[f'F{current_row}'] = 'Length (inches)'
#             ws[f'G{current_row}'] = 'Thickness (inches)'
#             ws[f'H{current_row}'] = 'Weight Per Unit'
#             ws[f'I{current_row}'] = 'Weight Total'
#             ws[f'J{current_row}'] = 'Cost Per Unit'
#             ws[f'K{current_row}'] = 'Cost Total'
#             ws[f'L{current_row}'] = 'Markup %'
#             ws[f'M{current_row}'] = 'Markup'
#             ws[f'N{current_row}'] = 'Burden'
#             ws[f'O{current_row}'] = 'Taxable 1 = Y 0 = N'
#             ws[f'P{current_row}'] = 'Tax Amount'
#             ws[f'Q{current_row}'] = 'Total'
#             ws[f'R{current_row}'] = 'Type'
#             current_row += 1
            
#             for sub in sub_items:
#                 ws[f'A{current_row}'] = sub.get('qty', 0)
#                 ws[f'B{current_row}'] = sub.get('unit', 'ea')
#                 ws[f'C{current_row}'] = sub.get('item_name', '')
#                 ws[f'D{current_row}'] = sub.get('description', '')
#                 ws[f'E{current_row}'] = f"{sub.get('width', 0.0):.3f}"
#                 ws[f'F{current_row}'] = f"{sub.get('length', 0.0):.3f}"
#                 ws[f'G{current_row}'] = f"{sub.get('thickness', 0.0):.3f}"
#                 ws[f'H{current_row}'] = sub.get('weight_per_unit', '') if sub.get('weight_per_unit') else ''
#                 ws[f'I{current_row}'] = f"{sub.get('weight_total', 0.0):.3f}"
#                 ws[f'J{current_row}'] = sub.get('cost_per_unit', 0.0)
#                 ws[f'K{current_row}'] = sub.get('cost_total', 0.0)
#                 ws[f'L{current_row}'] = sub.get('markup_pct', 0.0)
#                 ws[f'M{current_row}'] = sub.get('markup', 0.0)
#                 ws[f'N{current_row}'] = sub.get('burden', 0.0)
#                 ws[f'O{current_row}'] = sub.get('taxable', 0)
#                 ws[f'P{current_row}'] = sub.get('tax_amount', 0.0)
#                 ws[f'Q{current_row}'] = sub.get('total', 0.0)
#                 ws[f'R{current_row}'] = sub.get('type', 'Material')
                
#                 print(f"  └─ {sub.get('item_name', 'Unnamed sub-item')}: ${sub.get('total', 0.0):.2f}")
#                 current_row += 1
        
#         # Add blank row between items
#         current_row += 1
    
#     # Update summary cells (K2:Q2)
#     summary = data.get("summary", {})
#     if summary:
#         ws['K2'] = summary.get('cost_total', 0.0)
#         ws['L2'] = summary.get('markup_pct', 0.27)
#         ws['M2'] = summary.get('markup', 0.0)
#         ws['N2'] = summary.get('burden', 0.0)
#         ws['O2'] = summary.get('pct_with_burden', 0.36)
#         ws['P2'] = summary.get('tax_amount', 0.0)
#         ws['Q2'] = summary.get('sale_total', 0.0)
#         print(f"\n✓ Summary updated: Sale Total = ${summary.get('sale_total', 0.0):.2f}")
    
#     # Save the file
#     wb.save(OUTPUT_FILE)
#     print(f"\n✅ Quote worksheet saved to: {OUTPUT_FILE}")
#     print(f"📊 Total items added: {len(quote_items)}")
#     print(f"📍 Last row used: {current_row - 1}")

# def main():
#     # Load demo data
#     SCRIPT_DIR = Path(__file__).parent
#     ROOT_DIR = SCRIPT_DIR.parent
#     INPUT_DIR = ROOT_DIR / "input"
    
#     DEMO_DATA_FILE = INPUT_DIR / "quote_data.json"
    
#     if not DEMO_DATA_FILE.exists():
#         print(f"❌ Demo data file not found: {DEMO_DATA_FILE}")
#         return
    
#     # Load data
#     with open(DEMO_DATA_FILE, 'r', encoding='utf-8') as f:
#         data = json.load(f)
    
#     print(f"\n{'='*80}")
#     print("FILLING QUOTE WORKSHEET")
#     print('='*80)
    
#     fill_quote_worksheet(data)

# if __name__ == "__main__":
#     main()


#  new one
# import openpyxl
# from pathlib import Path
# import json
# from datetime import datetime
# import shutil

# def fill_quote_worksheet(data):
#     """
#     Fill the Quote Worksheet Excel file by updating existing template structure.
#     This preserves all buttons, formulas, and formatting.
#     """
#     SCRIPT_DIR = Path(__file__).parent
#     ROOT_DIR = SCRIPT_DIR.parent
#     TEMPLATE_DIR = ROOT_DIR / "templates"
#     OUTPUT_DIR = ROOT_DIR / "output"
    
#     TEMPLATE_FILE = TEMPLATE_DIR / "worksheet.xlsm"
#     timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
#     OUTPUT_FILE = OUTPUT_DIR / f"Filled_Quote_{timestamp}.xlsm"
    
#     OUTPUT_DIR.mkdir(exist_ok=True)
    
#     if not TEMPLATE_FILE.exists():
#         print(f"❌ Template file not found: {TEMPLATE_FILE}")
#         return
    
#     # Copy template to output first (preserves everything)
#     shutil.copy2(TEMPLATE_FILE, OUTPUT_FILE)
#     print(f"✅ Copied template to: {OUTPUT_FILE}")
    
#     # Now open and modify the copy
#     wb = openpyxl.load_workbook(OUTPUT_FILE, keep_vba=True)
    
#     # Select the correct sheet
#     if "Quote Worksheet" in wb.sheetnames:
#         ws = wb["Quote Worksheet"]
#     else:
#         print(f"⚠️  'Quote Worksheet' not found. Available sheets: {wb.sheetnames}")
#         ws = wb.active
    
#     print(f"📄 Working on sheet: {ws.title}")
    
#     # Update configuration section (D1-D8)
#     config = data.get("configuration", {})
#     ws['D1'] = config.get("tax_rate", 0.0825)
#     ws['D2'] = config.get("markup_labor", 0.00)
#     ws['D3'] = config.get("markup_other", 0.25)
#     ws['D4'] = config.get("lead_brick_price", 1.25)
#     ws['D5'] = config.get("lead_sheet_price", 1.75)
#     ws['D6'] = config.get("steel_price", 1.40)
#     ws['D7'] = config.get("poly_5_price", 850.00)
#     ws['D8'] = config.get("poly_hd_price", 350.00)
    
#     print("✓ Configuration updated")
    
#     # Clear existing item data (rows 18-23 in template)
#     # But keep the structure - just clear values
#     for row in range(18, 24):
#         for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
#             cell = ws[f'{col}{row}']
#             # Keep formatting but clear value
#             cell.value = None
    
#     print("✓ Cleared existing item data")
    
#     # Get quote items
#     quote_items = data.get("quote_items", [])
    
#     # Start at row 18 (where items begin in template)
#     current_row = 18
    
#     for idx, item in enumerate(quote_items, 1):
#         print(f"\n📦 Adding Item {idx}: {item.get('item_name', 'Unnamed')}")
        
#         # Check if this item has sub-items
#         sub_items = item.get('sub_items', [])
        
#         # Write main item row
#         ws[f'A{current_row}'] = item.get('qty', 1)
#         ws[f'B{current_row}'] = item.get('unit', 'ea')
#         ws[f'C{current_row}'] = item.get('item_name', '')
        
#         if not sub_items:
#             # For simple items without sub-items, write all data
#             ws[f'D{current_row}'] = item.get('description', '')
#             ws[f'E{current_row}'] = item.get('width', '')
#             ws[f'F{current_row}'] = item.get('length', '')
#             ws[f'G{current_row}'] = item.get('thickness', '')
#             ws[f'H{current_row}'] = item.get('weight_per_unit', '')
#             ws[f'I{current_row}'] = item.get('weight_total', '')
#             ws[f'J{current_row}'] = item.get('cost_per_unit', '')
#             ws[f'K{current_row}'] = item.get('cost_total', '')
#             ws[f'L{current_row}'] = item.get('markup_pct', '')
#             ws[f'M{current_row}'] = item.get('markup', '')
#             ws[f'N{current_row}'] = item.get('burden', '')
#             ws[f'O{current_row}'] = item.get('taxable', '')
#             ws[f'P{current_row}'] = item.get('tax_amount', '')
#             ws[f'Q{current_row}'] = item.get('total', '')
#             ws[f'R{current_row}'] = item.get('type', 'Material')
#             print(f"  ✓ Simple item: ${item.get('total', 0):.2f}")
#             current_row += 1
#         else:
#             # For items with sub-items (assemblies)
#             current_row += 1
            
#             # Add sub-item headers
#             ws[f'A{current_row}'] = 'Qty.'
#             ws[f'B{current_row}'] = 'Unit'
#             ws[f'C{current_row}'] = 'Item'
#             ws[f'D{current_row}'] = 'Description'
#             ws[f'E{current_row}'] = 'Width (inches)'
#             ws[f'F{current_row}'] = 'Length (inches)'
#             ws[f'G{current_row}'] = 'Thickness (inches)'
#             ws[f'H{current_row}'] = 'Weight Per Unit'
#             ws[f'I{current_row}'] = 'Weight Total'
#             ws[f'J{current_row}'] = 'Cost Per Unit'
#             ws[f'K{current_row}'] = 'Cost Total'
#             ws[f'L{current_row}'] = 'Markup %'
#             ws[f'M{current_row}'] = 'Markup'
#             ws[f'N{current_row}'] = 'Burden'
#             ws[f'O{current_row}'] = 'Taxable 1 = Y 0 = N'
#             ws[f'P{current_row}'] = 'Tax Amount'
#             ws[f'Q{current_row}'] = 'Total'
#             ws[f'R{current_row}'] = 'Type'
#             current_row += 1
            
#             # Add each sub-item
#             for sub in sub_items:
#                 ws[f'A{current_row}'] = sub.get('qty', '')
#                 ws[f'B{current_row}'] = sub.get('unit', '')
#                 ws[f'C{current_row}'] = sub.get('item_name', '')
#                 ws[f'D{current_row}'] = sub.get('description', '')
#                 ws[f'E{current_row}'] = sub.get('width', '')
#                 ws[f'F{current_row}'] = sub.get('length', '')
#                 ws[f'G{current_row}'] = sub.get('thickness', '')
#                 ws[f'H{current_row}'] = sub.get('weight_per_unit', '') if sub.get('weight_per_unit') else ''
#                 ws[f'I{current_row}'] = sub.get('weight_total', '')
#                 ws[f'J{current_row}'] = sub.get('cost_per_unit', '')
#                 ws[f'K{current_row}'] = sub.get('cost_total', '')
#                 ws[f'L{current_row}'] = sub.get('markup_pct', '')
#                 ws[f'M{current_row}'] = sub.get('markup', '')
#                 ws[f'N{current_row}'] = sub.get('burden', '')
#                 ws[f'O{current_row}'] = sub.get('taxable', '')
#                 ws[f'P{current_row}'] = sub.get('tax_amount', '')
#                 ws[f'Q{current_row}'] = sub.get('total', '')
#                 ws[f'R{current_row}'] = sub.get('type', '')
                
#                 print(f"  └─ {sub.get('item_name', '')}: ${sub.get('total', 0):.2f}")
#                 current_row += 1
        
#         # Add blank row between items (don't exceed row 24)
#         if current_row < 24:
#             current_row += 1
    
#     # Update summary cells (K2:Q2)
#     summary = data.get("summary", {})
#     if summary:
#         ws['K2'] = summary.get('cost_total', '')
#         ws['L2'] = summary.get('markup_pct', '')
#         ws['M2'] = summary.get('markup', '')
#         ws['N2'] = summary.get('burden', '')
#         ws['O2'] = summary.get('pct_with_burden', '')
#         ws['P2'] = summary.get('tax_amount', '')
#         ws['Q2'] = summary.get('sale_total', '')
#         print(f"\n✓ Summary updated: Sale Total = ${summary.get('sale_total', 0):.2f}")
    
#     # Save the file
#     wb.save(OUTPUT_FILE)
#     print(f"\n✅ Quote worksheet saved to: {OUTPUT_FILE}")
#     print(f"📊 Total items added: {len(quote_items)}")
#     print(f"🎯 All buttons and formatting preserved!")

# def main():
#     SCRIPT_DIR = Path(__file__).parent
#     ROOT_DIR = SCRIPT_DIR.parent
#     INPUT_DIR = ROOT_DIR / "input"
    
#     DEMO_DATA_FILE = INPUT_DIR / "quote_data.json"
    
#     if not DEMO_DATA_FILE.exists():
#         print(f"❌ Demo data file not found: {DEMO_DATA_FILE}")
#         return
    
#     with open(DEMO_DATA_FILE, 'r', encoding='utf-8') as f:
#         data = json.load(f)
    
#     print(f"\n{'='*80}")
#     print("FILLING QUOTE WORKSHEET - PRESERVING TEMPLATE STRUCTURE")
#     print('='*80)
    
#     fill_quote_worksheet(data)

# if __name__ == "__main__":
#     main()



import openpyxl
from pathlib import Path
import json
from datetime import datetime
import shutil

def fill_quote_worksheet(data):
    """
    Fill the Quote Worksheet Excel file by updating existing template structure.
    This preserves all buttons, formulas, and formatting.
    """
    SCRIPT_DIR = Path(__file__).parent
    ROOT_DIR = SCRIPT_DIR.parent
    TEMPLATE_DIR = ROOT_DIR / "templates"
    OUTPUT_DIR = ROOT_DIR / "output"
    
    TEMPLATE_FILE = TEMPLATE_DIR / "worksheet.xlsm"
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    OUTPUT_FILE = OUTPUT_DIR / f"Filled_Quote_{timestamp}.xlsm"
    
    OUTPUT_DIR.mkdir(exist_ok=True)
    
    if not TEMPLATE_FILE.exists():
        print(f"❌ Template file not found: {TEMPLATE_FILE}")
        return
    
    # Copy template to output first (preserves everything)
    shutil.copy2(TEMPLATE_FILE, OUTPUT_FILE)
    print(f"✅ Copied template to: {OUTPUT_FILE}")
    
    # Now open and modify the copy
    wb = openpyxl.load_workbook(OUTPUT_FILE, keep_vba=True)
    
    # Select the correct sheet
    if "Quote Worksheet" in wb.sheetnames:
        ws = wb["Quote Worksheet"]
    else:
        print(f"⚠️  'Quote Worksheet' not found. Available sheets: {wb.sheetnames}")
        ws = wb.active
    
    print(f"📄 Working on sheet: {ws.title}")
    
    # Update configuration section (D1-D8)
    config = data.get("configuration", {})
    ws['D1'] = config.get("tax_rate", 0.0825)
    ws['D2'] = config.get("markup_labor", 0.00)
    ws['D3'] = config.get("markup_other", 0.25)
    ws['D4'] = config.get("lead_brick_price", 1.25)
    ws['D5'] = config.get("lead_sheet_price", 1.75)
    ws['D6'] = config.get("steel_price", 1.40)
    ws['D7'] = config.get("poly_5_price", 850.00)
    ws['D8'] = config.get("poly_hd_price", 350.00)
    
    print("✓ Configuration updated")
    
    # Clear existing item data (rows 18-23 in template)
    # But keep the structure - just clear values
    for row in range(18, 24):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
            cell = ws[f'{col}{row}']
            # Keep formatting but clear value
            cell.value = None
    
    print("✓ Cleared existing item data")
    
    # Get quote items
    quote_items = data.get("quote_items", [])
    
    # Start at row 18 (where items begin in template)
    current_row = 18
    
    for idx, item in enumerate(quote_items, 1):
        print(f"\n📦 Adding Item {idx}: {item.get('item_name', 'Unnamed')}")
        
        # Check if this item has sub-items
        sub_items = item.get('sub_items', [])
        
        # Write main item row
        ws[f'A{current_row}'] = item.get('qty', 1)
        ws[f'B{current_row}'] = item.get('unit', 'ea')
        ws[f'C{current_row}'] = item.get('item_name', '')
        
        if not sub_items:
            # For simple items without sub-items, write all data
            ws[f'D{current_row}'] = item.get('description', '')
            ws[f'E{current_row}'] = item.get('width', '')
            ws[f'F{current_row}'] = item.get('length', '')
            ws[f'G{current_row}'] = item.get('thickness', '')
            ws[f'H{current_row}'] = item.get('weight_per_unit', '')
            ws[f'I{current_row}'] = item.get('weight_total', '')
            ws[f'J{current_row}'] = item.get('cost_per_unit', '')
            ws[f'K{current_row}'] = item.get('cost_total', '')
            ws[f'L{current_row}'] = item.get('markup_pct', '')
            ws[f'M{current_row}'] = item.get('markup', '')
            ws[f'N{current_row}'] = item.get('burden', '')
            ws[f'O{current_row}'] = item.get('taxable', '')
            ws[f'P{current_row}'] = item.get('tax_amount', '')
            ws[f'Q{current_row}'] = item.get('total', '')
            ws[f'R{current_row}'] = item.get('type', 'Material')
            print(f"  ✓ Simple item: ${item.get('total', 0):.2f}")
            current_row += 1
        else:
            # For items with sub-items (assemblies)
            current_row += 1
            
            # Add sub-item headers
            ws[f'A{current_row}'] = 'Qty.'
            ws[f'B{current_row}'] = 'Unit'
            ws[f'C{current_row}'] = 'Item'
            ws[f'D{current_row}'] = 'Description'
            ws[f'E{current_row}'] = 'Width (inches)'
            ws[f'F{current_row}'] = 'Length (inches)'
            ws[f'G{current_row}'] = 'Thickness (inches)'
            ws[f'H{current_row}'] = 'Weight Per Unit'
            ws[f'I{current_row}'] = 'Weight Total'
            ws[f'J{current_row}'] = 'Cost Per Unit'
            ws[f'K{current_row}'] = 'Cost Total'
            ws[f'L{current_row}'] = 'Markup %'
            ws[f'M{current_row}'] = 'Markup'
            ws[f'N{current_row}'] = 'Burden'
            ws[f'O{current_row}'] = 'Taxable 1 = Y 0 = N'
            ws[f'P{current_row}'] = 'Tax Amount'
            ws[f'Q{current_row}'] = 'Total'
            ws[f'R{current_row}'] = 'Type'
            current_row += 1
            
            # Add each sub-item
            for sub in sub_items:
                ws[f'A{current_row}'] = sub.get('qty', '')
                ws[f'B{current_row}'] = sub.get('unit', '')
                ws[f'C{current_row}'] = sub.get('item_name', '')
                ws[f'D{current_row}'] = sub.get('description', '')
                ws[f'E{current_row}'] = sub.get('width', '')
                ws[f'F{current_row}'] = sub.get('length', '')
                ws[f'G{current_row}'] = sub.get('thickness', '')
                ws[f'H{current_row}'] = sub.get('weight_per_unit', '') if sub.get('weight_per_unit') else ''
                ws[f'I{current_row}'] = sub.get('weight_total', '')
                ws[f'J{current_row}'] = sub.get('cost_per_unit', '')
                ws[f'K{current_row}'] = sub.get('cost_total', '')
                ws[f'L{current_row}'] = sub.get('markup_pct', '')
                ws[f'M{current_row}'] = sub.get('markup', '')
                ws[f'N{current_row}'] = sub.get('burden', '')
                ws[f'O{current_row}'] = sub.get('taxable', '')
                ws[f'P{current_row}'] = sub.get('tax_amount', '')
                ws[f'Q{current_row}'] = sub.get('total', '')
                ws[f'R{current_row}'] = sub.get('type', '')
                
                print(f"  └─ {sub.get('item_name', '')}: ${sub.get('total', 0):.2f}")
                current_row += 1
        
        # Add blank row between items (don't exceed row 24)
        if current_row < 24:
            current_row += 1
    
    # Update summary cells (K2:Q2)
    summary = data.get("summary", {})
    if summary:
        ws['K2'] = summary.get('cost_total', '')
        ws['L2'] = summary.get('markup_pct', '')
        ws['M2'] = summary.get('markup', '')
        ws['N2'] = summary.get('burden', '')
        ws['O2'] = summary.get('pct_with_burden', '')
        ws['P2'] = summary.get('tax_amount', '')
        ws['Q2'] = summary.get('sale_total', '')
        print(f"\n✓ Summary updated: Sale Total = ${summary.get('sale_total', 0):.2f}")
    
    # Save the file
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Quote worksheet saved to: {OUTPUT_FILE}")
    print(f"📊 Total items added: {len(quote_items)}")
    print(f"🎯 All buttons and formatting preserved!")

def main():
    SCRIPT_DIR = Path(__file__).parent
    ROOT_DIR = SCRIPT_DIR.parent
    INPUT_DIR = ROOT_DIR / "input"
    
    DEMO_DATA_FILE = INPUT_DIR / "quote_data2.json"
    
    if not DEMO_DATA_FILE.exists():
        print(f"❌ Demo data file not found: {DEMO_DATA_FILE}")
        return
    
    with open(DEMO_DATA_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    print(f"\n{'='*80}")
    print("FILLING QUOTE WORKSHEET - PRESERVING TEMPLATE STRUCTURE")
    print('='*80)
    
    fill_quote_worksheet(data)

if __name__ == "__main__":
    main()